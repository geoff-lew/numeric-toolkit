#!/usr/bin/env python3
"""
Close-checklist spreadsheet → Numeric Import workbook

Works on FloQast exports, BlackLine/Trintech extracts, or any homegrown xlsx/csv
whose rows are close tasks. Auto-detects column mappings from a synonym table;
explicit overrides win via --map.

Usage:
    python3 convert.py <input.xlsx|csv> [output.xlsx] [options]

Options:
    --sheet NAME          Sheet to read (xlsx only). Default: first sheet.
    --map FIELD=HEADER    Override auto-detected mapping. Repeatable.
                          FIELD ∈ category, task_name, frequency, preparer,
                                  preparer_deadline, reviewer,
                                  reviewer_deadline, tags
    --list-headers        Print the headers found in the file and exit.
    --dry-run             Detect columns, print the mapping, don't write output.

Examples:
    python3 convert.py FloQast_export.xlsx
    python3 convert.py blackline.xlsx out.xlsx --map task_name="Task Description"
    python3 convert.py homegrown.csv --map category="Process" --map preparer="Owner"
"""

from __future__ import annotations

import csv
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Optional

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write("openpyxl is required. Install with: pip install openpyxl\n")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Numeric-side constants (stable regardless of source tool)
# ---------------------------------------------------------------------------

TASK_NAME_MAX = 350
NAME_INLINE_MAX = 120  # short names go in Task Name only; longer also mirror to Description

# Business-day deadline guardrails. Anything outside these bounds is almost
# certainly a source-data error (typo, or a calendar day jammed into a BD
# column). Route those rows to Needs Review.
PREP_BD_MAX = 65
REV_BD_MAX = 45

# Weekly tasks get exploded into 4 monthly rows at BD+5/+10/+15/+20.
WEEKLY_SPLIT_BDS = [5, 10, 15, 20]

MONTH_TOKENS = {
    "jan", "january", "feb", "february", "mar", "march", "apr", "april",
    "may", "jun", "june", "jul", "july", "aug", "august", "sep", "sept",
    "september", "oct", "october", "nov", "november", "dec", "december",
}


# ---------------------------------------------------------------------------
# Column detection — synonyms for each Numeric-side field
# ---------------------------------------------------------------------------
#
# Order within each list matters for tie-breaking: the first synonym that
# matches a source header wins. Matching is case- and whitespace-insensitive
# (so "Preparer Email" matches "preparer_email", "PREPARER EMAIL", etc.).
#
# The goal is to cover the big three close tools (FloQast, BlackLine, Trintech)
# plus generic "Excel tracker" headers. If a header isn't in this list, the
# user can pass --map field="Exact Header".

FIELD_SYNONYMS: dict[str, list[str]] = {
    "category": [
        "FQ Folder", "Folder", "Group", "Section", "Category",
        "Area", "Process", "Process Area", "Cycle", "Workstream",
        "Close Area", "Module",
    ],
    "task_name": [
        "Description", "Task", "Task Name", "Task Description", "Name",
        "Activity", "Step", "Item", "Procedure", "Control Description",
        "Control", "Close Task",
    ],
    "frequency": [
        "Frequency", "Cadence", "Schedule", "Freq", "Recurrence",
    ],
    "preparer": [
        "Preparer", "Preparer Email", "Assignee", "Owner",
        "Prepared By", "Responsible", "Assigned To", "Performer",
    ],
    "preparer_deadline": [
        "Preparer Deadline", "Prep Due", "Prep BD", "Preparer Due",
        "Preparer Business Day", "Due Date", "Due BD", "Business Day",
        "BD", "Day", "Offset",
    ],
    "reviewer": [
        "Reviewer", "Reviewer Email", "Approver", "Review By",
        "Reviewed By", "Secondary Reviewer",
    ],
    "reviewer_deadline": [
        "Reviewer Deadline", "Review Due", "Rev BD", "Reviewer Due",
        "Reviewer Business Day", "Approval Due",
    ],
    "tags": [
        "Tags", "Labels", "Notes", "Remarks",
    ],
}

REQUIRED_FIELDS = ("category", "task_name", "frequency")


# ---------------------------------------------------------------------------
# Text helpers
# ---------------------------------------------------------------------------

def _norm(s: str) -> str:
    """Normalize a header for comparison: lowercase, collapse whitespace/underscores."""
    if s is None:
        return ""
    return re.sub(r"[\s_]+", " ", str(s).strip().lower())


def _split_multi(raw: Optional[str]) -> list[str]:
    """Split a multi-assignee cell on ; or , and clean up."""
    if not raw:
        return []
    parts = re.split(r"[;,]", str(raw))
    return [p.strip() for p in parts if p and p.strip()]


def _parse_tags(raw: Optional[str]) -> list[str]:
    if not raw:
        return []
    parts = re.split(r"[;,]", str(raw))
    out = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        if not p.startswith("#"):
            p = "#" + p
        out.append(p)
    return out


def _has_je_tag(tags: list[str]) -> bool:
    lowered = {t.lower() for t in tags}
    return bool(lowered & {"#je", "#journal_entry", "#journalentry"})


def _is_weekly(raw: Optional[str]) -> bool:
    if not raw:
        return False
    return str(raw).strip().lower() in ("weekly", "week")


def _map_frequency(raw: Optional[str]) -> tuple[str, bool, str]:
    """Returns (numeric_freq, was_remapped, original_value)."""
    if not raw or not str(raw).strip():
        return "Monthly", True, ""

    original = str(raw).strip()
    val = original.lower()

    if val == "monthly":
        return "Monthly", False, original
    if val == "quarterly":
        return "Quarterly", False, original
    if val in ("annual", "annually", "yearly", "yearly once"):
        return "Annually", val != "annually", original
    if val in ("weekly", "week"):
        return "Monthly", True, original  # safety net; normally pre-handled

    if val.startswith("custom"):
        tokens = re.findall(r"[a-z]+", val)
        months = {t for t in tokens if t in MONTH_TOKENS}
        n = len(months)
        if n >= 6:
            return "Monthly", True, original
        if 2 <= n <= 5:
            return "Quarterly", True, original
        if n == 1:
            return "Annually", True, original
        return "Monthly", True, original

    return "Monthly", True, original


def _int_or_blank(raw) -> str:
    """Coerce a single deadline cell to a clean integer string, or blank.

    Close tools often decorate deadlines with trailing '*' (weekend/holiday
    marker) that isn't meaningful to Numeric — strip it.
    """
    if raw is None or raw == "":
        return ""
    s = str(raw).strip().rstrip("*").strip()
    if not s:
        return ""
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
        return str(f)
    except (TypeError, ValueError):
        return s


def _split_deadlines(raw) -> list[str]:
    """Split a deadline cell that may contain multiple values (e.g. '3; 5')."""
    if raw is None or str(raw).strip() == "":
        return []
    parts = re.split(r"[;,]", str(raw))
    return [_int_or_blank(p) for p in parts if p is not None and str(p).strip() != ""]


def _strip_category_prefix(folder: str) -> str:
    """Folders like '01 Cash and Cash Equivalents' → 'Cash and Cash Equivalents'.

    Only strip when the digits are followed by a real separator — otherwise
    '401k Admin' would wrongly become 'k Admin'. Requires one-or-more of
    whitespace / . / - / : after the leading digits.
    """
    if not folder:
        return folder
    stripped = re.sub(r"^\s*\d+[\s.\-:]+", "", folder).strip()
    return stripped or folder


def _clean_cell(raw) -> str:
    if raw is None:
        return ""
    return str(raw).strip()


# ---------------------------------------------------------------------------
# Input reading (xlsx and csv)
# ---------------------------------------------------------------------------

def read_spreadsheet(path: Path, sheet: Optional[str] = None) -> tuple[list[str], list[list]]:
    """Return (headers, rows) where rows is a list of lists of cell values."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = [list(r) for r in reader]
    elif suffix in (".xlsx", ".xlsm"):
        wb = load_workbook(path, data_only=True, read_only=True)
        if sheet:
            if sheet not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
            ws = wb[sheet]
        else:
            ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
    else:
        raise ValueError(f"Unsupported file type: {suffix} (need .xlsx, .xlsm, or .csv)")

    if not rows:
        raise ValueError("File is empty")

    # Drop leading blank rows (some exports have title banners before the header).
    while rows and all(c is None or str(c).strip() == "" for c in rows[0]):
        rows.pop(0)

    if not rows:
        raise ValueError("No header row found")

    headers = [_clean_cell(c) for c in rows[0]]
    body = rows[1:]
    return headers, body


def detect_columns(
    headers: list[str], overrides: dict[str, str]
) -> tuple[dict[str, Optional[int]], dict[str, str]]:
    """Return (field → column index, field → source header used).

    overrides is {field: exact_header_string}. These win over auto-detection.
    Missing fields map to None.
    """
    normalized = {_norm(h): i for i, h in enumerate(headers) if h}
    resolved_idx: dict[str, Optional[int]] = {}
    resolved_name: dict[str, str] = {}

    for field, synonyms in FIELD_SYNONYMS.items():
        # Explicit override?
        if field in overrides:
            target = _norm(overrides[field])
            if target not in normalized:
                raise ValueError(
                    f"--map {field}={overrides[field]!r}: that header isn't in the file. "
                    f"Available headers: {headers}"
                )
            resolved_idx[field] = normalized[target]
            resolved_name[field] = headers[normalized[target]]
            continue

        # Auto-detect: first synonym that's present wins.
        for syn in synonyms:
            key = _norm(syn)
            if key in normalized:
                resolved_idx[field] = normalized[key]
                resolved_name[field] = headers[normalized[key]]
                break
        else:
            resolved_idx[field] = None

    return resolved_idx, resolved_name


def build_raw_rows(
    body: list[list], col_idx: dict[str, Optional[int]]
) -> list[dict]:
    """Project each body row into a dict keyed by our canonical field names."""
    out = []
    for row in body:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        def get(field: str):
            i = col_idx.get(field)
            if i is None or i >= len(row):
                return None
            return row[i]

        out.append({
            "folder": _clean_cell(get("category")),
            "description": _clean_cell(get("task_name")),
            "frequency": _clean_cell(get("frequency")),
            "preparer": _clean_cell(get("preparer")),
            "preparer_deadline": get("preparer_deadline"),
            "reviewer": _clean_cell(get("reviewer")),
            "reviewer_deadline": get("reviewer_deadline"),
            "tags": _clean_cell(get("tags")),
        })
    return out


# ---------------------------------------------------------------------------
# Flags & row building
# ---------------------------------------------------------------------------

HARD_ISSUE_PREFIXES = (
    "missing Description",
    "missing Preparer",
    "Task Name truncated",
    "Frequency 'Custom",
    "Frequency '(blank)'",
    "Frequency 'Unknown",
    "Preparer due out of range",
    "Reviewer due out of range",
    "Preparer due is zero",
    "Reviewer due is zero",
)


def _is_hard(flag: dict) -> bool:
    issue = flag["issue"]
    return any(issue.startswith(p) for p in HARD_ISSUE_PREFIXES)


def build_rows(raw_rows: list[dict]) -> tuple[list[dict], list[dict], list[str], list[str], list[dict], int]:
    """Main classification/normalization step.

    Returns (main_rows, unprocessed_rows, prep_cols, rev_cols, flags, dropped_extras).
    Numeric only supports one preparer and one reviewer per task, so we keep
    the first and count drops for transparency.
    """
    prep_cols = ["Preparer"]
    prep_due_cols = ["Prep. Due"]
    rev_cols = ["Reviewer"]
    rev_due_cols = ["Rev. Due"]
    dropped_extras = 0

    flags = []
    main = []
    unprocessed = []

    def _validate_bds(bds: list, limit: int, label: str, src_row: int) -> list[dict]:
        """Coerce BD=0 to BD=1 (Numeric rejects BD=0) and hard-flag values
        outside ±limit. Mutates ``bds`` in place so the coerced value flows
        through to the emitted row."""
        out_flags = []
        for idx, v in enumerate(bds):
            if v == "" or v is None:
                continue
            try:
                n = int(float(str(v)))
            except (TypeError, ValueError):
                continue
            if n == 0:
                # Numeric treats BD=0 as invalid — shift to BD=1 silently.
                bds[idx] = "1"
                continue
            if abs(n) > limit:
                out_flags.append({
                    "row": src_row,
                    "issue": f"{label} due out of range (BD={n}, limit ±{limit})",
                    "value": n,
                })
        return out_flags

    for i, r in enumerate(raw_rows, start=2):  # +2: header is row 1 in source
        row_flags = []
        desc = r["description"]
        folder_raw = r["folder"] or "Uncategorized"
        folder = _strip_category_prefix(folder_raw)

        weekly = _is_weekly(r["frequency"])

        # Task Name & Description
        if not desc:
            task_name_base = f"[untitled task from {folder}]"
            description = ""
            row_flags.append({"row": i, "issue": "missing Description in source", "value": ""})
        elif len(desc) <= NAME_INLINE_MAX:
            task_name_base = desc
            description = ""
        elif len(desc) <= TASK_NAME_MAX:
            task_name_base = desc
            description = desc
        else:
            task_name_base = desc[: TASK_NAME_MAX - 3] + "..."
            description = desc
            row_flags.append({
                "row": i,
                "issue": f"Task Name truncated (source was {len(desc)} chars)",
                "value": desc[:80] + "…",
            })

        # Frequency
        if weekly:
            freq = "Monthly"
        else:
            freq, remapped, orig_freq = _map_frequency(r["frequency"])
            if remapped:
                label = orig_freq if orig_freq else "(blank)"
                row_flags.append({"row": i, "issue": f"Frequency '{label}' → {freq}", "value": orig_freq})

        # Type (from tags)
        tags = _parse_tags(r["tags"])
        task_type = "Journal Entry" if _has_je_tag(tags) else "Task"

        # Assignees — collapse to single slot
        prep_all = _split_multi(r["preparer"])
        rev_all = _split_multi(r["reviewer"])
        if len(prep_all) > 1 or len(rev_all) > 1:
            dropped_extras += 1
        prep_list = prep_all[:1]
        rev_list = rev_all[:1]
        if not prep_list:
            row_flags.append({"row": i, "issue": "missing Preparer", "value": ""})

        # Deadlines
        prep_due_list = _split_deadlines(r["preparer_deadline"])[:1]
        rev_due_list = _split_deadlines(r["reviewer_deadline"])[:1]

        # Weekly explode or single emit
        if weekly:
            gap = 1
            if prep_due_list and rev_due_list:
                try:
                    gap = int(float(str(rev_due_list[0]))) - int(float(str(prep_due_list[0])))
                except (TypeError, ValueError):
                    gap = 1
            emissions = []
            for wk_idx, prep_bd in enumerate(WEEKLY_SPLIT_BDS, start=1):
                name = f"{task_name_base} - Week {wk_idx}"
                wk_prep = [str(prep_bd)] if prep_list else []
                wk_rev = [str(prep_bd + gap)] if rev_list else []
                emissions.append((name, wk_prep, wk_rev))
        else:
            emissions = [(task_name_base, prep_due_list, rev_due_list)]

        for name_out, prep_bds_out, rev_bds_out in emissions:
            emit_flags = list(row_flags)
            # If there's no preparer/reviewer, there can't be a due date
            # for that slot — clear it so the output is consistent.
            if not prep_list:
                prep_bds_out = []
            if not rev_list:
                rev_bds_out = []
            emit_flags += _validate_bds(prep_bds_out, PREP_BD_MAX, "Preparer", i)
            emit_flags += _validate_bds(rev_bds_out, REV_BD_MAX, "Reviewer", i)

            row_out = {
                "Category": folder,
                "Task Name": name_out,
                "Frequency": freq,
                "Type": task_type,
                "Description": description,
            }
            for j, col in enumerate(prep_cols):
                row_out[col] = prep_list[j] if j < len(prep_list) else ""
            for j, col in enumerate(rev_cols):
                row_out[col] = rev_list[j] if j < len(rev_list) else ""
            for j, col in enumerate(prep_due_cols):
                row_out[col] = prep_bds_out[j] if j < len(prep_bds_out) else ""
            for j, col in enumerate(rev_due_cols):
                row_out[col] = rev_bds_out[j] if j < len(rev_bds_out) else ""

            flags.extend(emit_flags)
            hard_flags = [f for f in emit_flags if _is_hard(f)]
            if hard_flags:
                row_out["_issues"] = " | ".join(f["issue"] for f in hard_flags)
                row_out["_source_row"] = i
                unprocessed.append(row_out)
            else:
                main.append(row_out)

    return main, unprocessed, prep_cols, rev_cols, flags, dropped_extras


def _build_headers(prep_cols, rev_cols):
    prep_due = ["Prep. Due"] + [f"Prep. {i+1} Due" for i in range(1, len(prep_cols))]
    rev_due = ["Rev. Due"] + [f"Rev. {i+1} Due" for i in range(1, len(rev_cols))]
    return (
        ["Category", "Task Name", "Frequency", "Type", "Description"]
        + prep_cols + rev_cols
        + prep_due + rev_due
    )


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def write_xlsx(main_rows, unprocessed_rows, prep_cols, rev_cols, out_path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    headers = _build_headers(prep_cols, rev_cols)

    ws1 = wb.active
    ws1.title = "Numeric Import"
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(vertical="center")
    for r in main_rows:
        ws1.append([r.get(h, "") for h in headers])
    ws1.freeze_panes = "A2"
    for col_idx, h in enumerate(headers, start=1):
        ws1.column_dimensions[ws1.cell(row=1, column=col_idx).column_letter].width = min(max(12, len(h) + 2), 40)

    ws2 = wb.create_sheet("Needs Review")
    review_headers = ["Source Row", "Issues"] + headers
    ws2.append(review_headers)
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.alignment = Alignment(vertical="center")
    for r in unprocessed_rows:
        row_vals = [r.get("_source_row", ""), r.get("_issues", "")] + [r.get(h, "") for h in headers]
        ws2.append(row_vals)
    ws2.freeze_panes = "C2"
    for col_idx, h in enumerate(review_headers, start=1):
        ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = min(max(12, len(h) + 2), 50)

    wb.save(out_path)


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------

def print_mapping(resolved_name: dict[str, str], headers: list[str]):
    """Show the user which source header was used for each Numeric field."""
    print("Column mapping:")
    all_fields = list(FIELD_SYNONYMS.keys())
    for field in all_fields:
        src = resolved_name.get(field)
        if src:
            print(f"  {field:<20} ← {src!r}")
        else:
            tag = "REQUIRED" if field in REQUIRED_FIELDS else "optional"
            print(f"  {field:<20} ← (not found, {tag})")
    # Note unused headers for transparency
    used = {_norm(v) for v in resolved_name.values() if v}
    unused = [h for h in headers if h and _norm(h) not in used]
    if unused:
        print(f"  Ignored headers: {unused}")


def print_summary(main_rows, unprocessed_rows, flags, weekly_source_count, dropped_extras, xlsx_path):
    type_counts = Counter(r["Type"] for r in main_rows + unprocessed_rows)
    freq_counts = Counter(r["Frequency"] for r in main_rows)

    print(f"\n✓ {len(main_rows) + len(unprocessed_rows)} output rows "
          f"(weekly source tasks expanded: {weekly_source_count} → {weekly_source_count * 4})")
    print(f"  → {len(main_rows)} clean rows on 'Numeric Import' tab")
    print(f"  → {len(unprocessed_rows)} rows routed to 'Needs Review' tab")
    if dropped_extras:
        print(f"  ⚠ {dropped_extras} rows had 2+ preparers/reviewers — kept first only "
              f"(duplicate the task in Numeric post-import if dual sign-off is required)")
    print()
    if type_counts:
        print(f"  Type:          " + "  |  ".join(f"{k} {v}" for k, v in type_counts.most_common()))
    if freq_counts:
        print(f"  Frequency:     " + "  |  ".join(f"{k} {v}" for k, v in freq_counts.most_common()))

    if unprocessed_rows:
        print(f"\n  Needs Review breakdown:")
        by_issue = Counter()
        for r in unprocessed_rows:
            for piece in (r["_issues"] or "").split(" | "):
                piece = piece.strip()
                if piece:
                    if piece.startswith("Frequency"):
                        by_issue["Frequency needs review"] += 1
                    elif piece.startswith("Task Name truncated"):
                        by_issue["Task Name truncated"] += 1
                    elif piece.startswith("Preparer due out of range"):
                        by_issue["Preparer due out of range"] += 1
                    elif piece.startswith("Reviewer due out of range"):
                        by_issue["Reviewer due out of range"] += 1
                    elif piece.startswith("Preparer due is zero"):
                        by_issue["Preparer due is zero"] += 1
                    elif piece.startswith("Reviewer due is zero"):
                        by_issue["Reviewer due is zero"] += 1
                    else:
                        by_issue[piece] += 1
        for issue, n in by_issue.most_common():
            print(f"    {n:3}  {issue}")

    print(f"\n  Output: {xlsx_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _parse_args(argv):
    if len(argv) < 2:
        print(__doc__)
        sys.exit(1)

    in_path = None
    out_path = None
    sheet = None
    overrides: dict[str, str] = {}
    list_headers = False
    dry_run = False

    i = 1
    positional = []
    while i < len(argv):
        a = argv[i]
        if a == "--sheet" and i + 1 < len(argv):
            sheet = argv[i + 1]
            i += 2
        elif a == "--map" and i + 1 < len(argv):
            pair = argv[i + 1]
            if "=" not in pair:
                print(f"--map requires FIELD=HEADER, got: {pair!r}", file=sys.stderr)
                sys.exit(1)
            field, header = pair.split("=", 1)
            field = field.strip()
            if field not in FIELD_SYNONYMS:
                print(
                    f"Unknown --map field {field!r}. Valid: {list(FIELD_SYNONYMS)}",
                    file=sys.stderr,
                )
                sys.exit(1)
            overrides[field] = header.strip()
            i += 2
        elif a == "--list-headers":
            list_headers = True
            i += 1
        elif a == "--dry-run":
            dry_run = True
            i += 1
        elif a == "-h" or a == "--help":
            print(__doc__)
            sys.exit(0)
        elif a.startswith("--"):
            print(f"Unknown option: {a}", file=sys.stderr)
            sys.exit(1)
        else:
            positional.append(a)
            i += 1

    if not positional:
        print(__doc__)
        sys.exit(1)

    in_path = Path(positional[0])
    if len(positional) >= 2:
        out_path = Path(positional[1])

    return in_path, out_path, sheet, overrides, list_headers, dry_run


def main(argv):
    in_path, out_path, sheet, overrides, list_headers, dry_run = _parse_args(argv)

    if not in_path.exists():
        print(f"Input file not found: {in_path}", file=sys.stderr)
        sys.exit(1)

    headers, body = read_spreadsheet(in_path, sheet=sheet)

    if list_headers:
        print(f"Headers in {in_path.name}:")
        for h in headers:
            print(f"  {h!r}")
        return

    try:
        col_idx, resolved_name = detect_columns(headers, overrides)
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

    print_mapping(resolved_name, headers)
    print()

    # Required fields
    missing_required = [f for f in REQUIRED_FIELDS if col_idx.get(f) is None]
    if missing_required:
        print(
            f"\n❌ Could not find required columns: {missing_required}",
            file=sys.stderr,
        )
        print(
            f"   Use --map to specify them explicitly, e.g. "
            f"--map {missing_required[0]}=\"<Your Header>\"",
            file=sys.stderr,
        )
        sys.exit(2)

    if dry_run:
        print("(dry-run: no file written)")
        return

    if out_path is None:
        out_path = in_path.parent / (in_path.stem + "_Numeric_Import.xlsx")
    if out_path.suffix.lower() != ".xlsx":
        out_path = out_path.with_suffix(".xlsx")

    raw_rows = build_raw_rows(body, col_idx)
    weekly_source_count = sum(1 for r in raw_rows if _is_weekly(r["frequency"]))
    main_rows, unprocessed, prep_cols, rev_cols, flags, dropped_extras = build_rows(raw_rows)
    write_xlsx(main_rows, unprocessed, prep_cols, rev_cols, out_path)
    print_summary(main_rows, unprocessed, flags, weekly_source_count, dropped_extras, out_path)


if __name__ == "__main__":
    main(sys.argv)
