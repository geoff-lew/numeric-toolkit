#!/usr/bin/env python3
"""
Build the audit evidence Excel workbook from extracted Numeric data.

Usage:
    python build_workbook.py \
        --workspace-name "WHOOP" \
        --period-slug "Dec 2025" \
        --period-status "open" \
        --rec-tasks rec_tasks.json \
        --checklist-tasks checklist_tasks.json \
        --events all_events.json \
        --comments all_comments.json \
        --users user_map.json \
        --output output.xlsx

All input files are JSON. See the docstrings below for expected schemas.
"""

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Styles ──────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Arial", size=9)
TITLE_FONT = lambda color: Font(name="Arial", bold=True, size=14, color=color)
SUBTITLE_FONT = Font(name="Arial", size=10, italic=True, color="808080")
SECTION_FONT = Font(name="Arial", bold=True, size=10, color="2F5496")
BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
COMPLETE_FILL = PatternFill("solid", fgColor="E2EFDA")
PENDING_FILL = PatternFill("solid", fgColor="FCE4D6")
SKIP_FILL = PatternFill("solid", fgColor="F2F2F2")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FILLS = {
    "blue": PatternFill("solid", fgColor="2F5496"),
    "orange": PatternFill("solid", fgColor="C55A11"),
    "green": PatternFill("solid", fgColor="548235"),
    "purple": PatternFill("solid", fgColor="7030A0"),
    "gold": PatternFill("solid", fgColor="BF8F00"),
}


def style_cell(cell, font=DATA_FONT, fill=None, align=None):
    cell.font = font
    cell.border = BORDER
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align


def status_fill(val):
    if val == "COMPLETE":
        return COMPLETE_FILL
    elif val == "PENDING":
        return PENDING_FILL
    elif val in ("SKIPPED", "IMMATERIAL"):
        return SKIP_FILL
    return None


def strip_html(text):
    if not text:
        return ""
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</p>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def fmt_timestamp(iso_str):
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return iso_str


def write_header_row(ws, row, headers, fill_key):
    fill = HEADER_FILLS[fill_key]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def write_title(ws, title, subtitle, color, merge_end="I"):
    ws.merge_cells(f"A1:{merge_end}1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT(color)
    ws.merge_cells(f"A2:{merge_end}2")
    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT


# ── Sheet Builders ──────────────────────────────────────────────────────────

def build_task_sheet(wb, sheet_name, tab_color, fill_key, title, subtitle,
                     tasks, user_map, headers, col_widths, include_link=False):
    ws = wb.create_sheet(sheet_name) if sheet_name != wb.sheetnames[0] else wb.active
    if sheet_name == wb.sheetnames[0]:
        ws.title = sheet_name
    ws.sheet_properties.tabColor = tab_color

    merge_end = get_column_letter(len(headers))
    write_title(ws, title, subtitle, tab_color, merge_end)
    write_header_row(ws, 4, headers, fill_key)

    for i, task in enumerate(tasks):
        row = i + 5
        prep = user_map.get(task.get("prep_assignee", ""), task.get("prep_assignee", ""))
        rev = user_map.get(task.get("review_assignee", ""), task.get("review_assignee", ""))
        if not rev or rev == task.get("review_assignee", ""):
            rev = "" if not task.get("review_assignee") else rev

        vals = [
            task.get("name", ""),
            prep,
            task.get("prep_status", ""),
            task.get("prep_due", ""),
            rev,
            task.get("review_status", ""),
            task.get("review_due", ""),
        ]
        if include_link:
            vals.append(task.get("url", ""))

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            sf = status_fill(val)
            style_cell(cell, fill=sf)

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    last_row = 4 + len(tasks)
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{last_row}"
    return ws


def build_activity_log(wb, events, user_map, workspace_name, period_slug):
    ws = wb.create_sheet("Activity Log")
    ws.sheet_properties.tabColor = "548235"

    meaningful = [e for e in events if e.get("action_key") != "system_create_task"]
    meaningful.sort(key=lambda x: x.get("occurred_at", ""))

    write_title(
        ws,
        f"{workspace_name} - {period_slug} Close: Complete Activity Log",
        f"Generated: {datetime.now().strftime('%B %d, %Y')} | {len(meaningful)} events (excl. system-generated)",
        "548235",
        "F",
    )

    headers = ["Date/Time (UTC)", "Event", "Action By", "Status Change", "Task ID", "Via"]
    write_header_row(ws, 4, headers, "green")

    for i, evt in enumerate(meaningful):
        row = i + 5
        user = user_map.get(evt.get("by_user"), evt.get("by_user", "System"))
        sc = ""
        if evt.get("status_changed_from") or evt.get("status_changed_to"):
            sc = f'{evt.get("status_changed_from", "-")} → {evt.get("status_changed_to", "-")}'

        vals = [
            fmt_timestamp(evt.get("occurred_at", "")),
            evt.get("event", ""),
            user,
            sc,
            evt.get("task_id", ""),
            evt.get("via", "") or "",
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            action = evt.get("action_key", "")
            fill = None
            if action == "approve_task_review":
                fill = COMPLETE_FILL
            elif action == "return_task_review":
                fill = YELLOW_FILL
            style_cell(cell, fill=fill)

    widths = [22, 35, 25, 22, 40, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:F{4 + len(meaningful)}"
    return ws, len(meaningful)


def build_comments_sheet(wb, comments, user_map, workspace_name, period_slug):
    """
    comments: list of dicts, each with keys:
        task_name, task_type ("Reconciliation" or "Checklist"),
        body, created_at, user_id, review_note (bool),
        resolved (bool), resolved_by, resolved_at
    """
    ws = wb.create_sheet("Review Notes & Comments")
    ws.sheet_properties.tabColor = "7030A0"

    write_title(
        ws,
        f"{workspace_name} - {period_slug} Close: Review Notes & Comments",
        f"Generated: {datetime.now().strftime('%B %d, %Y')} | {len(comments)} total entries",
        "7030A0",
        "I",
    )

    headers = [
        "Task Name", "Task Type", "Comment / Note", "Author", "Date",
        "Review Note?", "Resolved?", "Resolved By", "Resolved Date",
    ]
    write_header_row(ws, 4, headers, "purple")

    # Sort by task name then date
    comments.sort(key=lambda c: (c.get("task_name", ""), c.get("created_at", "")))

    for i, c in enumerate(comments):
        row = i + 5
        is_rn = c.get("review_note", False)
        resolved = c.get("resolved", False)
        resolved_by = c.get("resolved_by", "") or ""
        if resolved_by and resolved_by.startswith("usr_"):
            resolved_by = user_map.get(resolved_by, resolved_by)
        author = c.get("author", "") or ""
        if not author or author.startswith("usr_"):
            author = user_map.get(c.get("user_id", ""), c.get("user_id", ""))

        vals = [
            c.get("task_name", ""),
            c.get("task_type", ""),
            strip_html(c.get("body", "")),
            author,
            fmt_timestamp(c.get("created_at", "")),
            "Yes" if is_rn else "No",
            "Yes" if resolved else ("No" if is_rn else ""),
            resolved_by,
            fmt_timestamp(c.get("resolved_at", "")) if c.get("resolved_at") else "",
        ]

        # Determine row fill
        row_fill = None
        if is_rn and not resolved:
            row_fill = YELLOW_FILL
        elif is_rn and resolved:
            row_fill = COMPLETE_FILL

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            style_cell(cell, fill=row_fill)
            if col == 3:  # comment body — wrap text
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [45, 16, 60, 22, 20, 14, 12, 22, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:I{4 + len(comments)}"
    return ws


def build_summary(wb, rec_tasks, checklist_tasks, events, comments, user_map,
                  workspace_name, period_slug):
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "BF8F00"

    ws.merge_cells("A1:D1")
    ws["A1"] = f"{workspace_name} - {period_slug} Close: Summary Statistics"
    ws["A1"].font = TITLE_FONT("BF8F00")

    row = 3

    # ── Reconciliation status ──
    ws.cell(row=row, column=1, value="Reconciliation Tasks").font = SECTION_FONT
    row += 1
    for label, key in [("Preparer", "prep_status"), ("Reviewer", "review_status")]:
        counts = Counter(t.get(key, "") for t in rec_tasks)
        for status, count in sorted(counts.items()):
            ws.cell(row=row, column=1, value=f"  {label} - {status or 'Unset'}").font = DATA_FONT
            ws.cell(row=row, column=2, value=count).font = DATA_FONT
            row += 1

    row += 1
    ws.cell(row=row, column=1, value="Checklist Tasks").font = SECTION_FONT
    row += 1
    cl_counts = Counter(t.get("prep_status", "") for t in checklist_tasks)
    for status, count in sorted(cl_counts.items()):
        ws.cell(row=row, column=1, value=f"  Preparer - {status or 'Unset'}").font = DATA_FONT
        ws.cell(row=row, column=2, value=count).font = DATA_FONT
        row += 1

    # ── Event breakdown ──
    row += 1
    ws.cell(row=row, column=1, value="Event Type Breakdown").font = SECTION_FONT
    row += 1
    action_labels = {
        "approve_task_review": "Reviews Approved",
        "submit_task": "Tasks Submitted / Completed",
        "return_task_review": "Reviews Returned / Rejected",
        "auto_unsubmit_preparer_task": "Auto-Unsubmitted (Preparer)",
        "auto_submit_reviewer_task": "Auto-Submitted (Reviewer)",
        "auto_submit_preparer_task": "Auto-Submitted (Preparer)",
        "system_create_task": "Tasks Generated from Template",
        "resolve_review_note": "Review Notes Resolved",
        "unsubmit_task": "Tasks Unsubmitted",
        "assign_task": "Task Assignments",
        "edit_target_date": "Due Date Changes",
        "create_task": "Tasks Created Manually",
    }
    action_counts = Counter(e.get("action_key", "") for e in events)
    for action, count in sorted(action_counts.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=f"  {action_labels.get(action, action)}").font = DATA_FONT
        ws.cell(row=row, column=2, value=count).font = DATA_FONT
        row += 1

    # ── Activity by user ──
    row += 1
    ws.cell(row=row, column=1, value="Activity by User").font = SECTION_FONT
    row += 1
    user_actions = Counter(
        user_map.get(e.get("by_user"), "System")
        for e in events
        if e.get("action_key") != "system_create_task"
    )
    for user, count in sorted(user_actions.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=f"  {user}").font = DATA_FONT
        ws.cell(row=row, column=2, value=count).font = DATA_FONT
        row += 1

    # ── Review notes summary ──
    row += 1
    ws.cell(row=row, column=1, value="Review Notes & Comments").font = SECTION_FONT
    row += 1
    total_comments = len(comments)
    review_notes = [c for c in comments if c.get("review_note")]
    resolved = sum(1 for c in review_notes if c.get("resolved"))
    unresolved = len(review_notes) - resolved
    general_comments = total_comments - len(review_notes)

    for label, val in [
        ("Total comments & notes", total_comments),
        ("Review notes (formal)", len(review_notes)),
        ("  Resolved", resolved),
        ("  Unresolved", unresolved),
        ("General comments", general_comments),
    ]:
        ws.cell(row=row, column=1, value=f"  {label}").font = DATA_FONT
        ws.cell(row=row, column=2, value=val).font = DATA_FONT
        row += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 15
    return ws


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Build audit evidence Excel workbook")
    parser.add_argument("--workspace-name", required=True)
    parser.add_argument("--period-slug", required=True)
    parser.add_argument("--period-status", default="open")
    parser.add_argument("--rec-tasks", required=True, help="JSON file of reconciliation tasks")
    parser.add_argument("--checklist-tasks", required=True, help="JSON file of checklist tasks")
    parser.add_argument("--events", required=True, help="JSON file of all events")
    parser.add_argument("--comments", required=True, help="JSON file of all comments")
    parser.add_argument("--users", required=True, help="JSON file mapping user_id → name")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    args = parser.parse_args()

    with open(args.rec_tasks) as f:
        rec_tasks = json.load(f)
    with open(args.checklist_tasks) as f:
        checklist_tasks = json.load(f)
    with open(args.events) as f:
        events = json.load(f)
    with open(args.comments) as f:
        comments = json.load(f)
    with open(args.users) as f:
        user_map = json.load(f)

    now = datetime.now().strftime("%B %d, %Y")
    ws_name = args.workspace_name
    period = args.period_slug
    status = args.period_status

    wb = Workbook()

    # Sheet 1: Reconciliation Tasks
    rec_headers = ["Account Name", "Preparer", "Prep Status", "Prep Due",
                   "Reviewer", "Review Status", "Review Due", "Link"]
    rec_widths = [55, 25, 14, 12, 25, 14, 12, 65]
    ws1 = wb.active
    ws1.title = "Reconciliation Tasks"
    ws1.sheet_properties.tabColor = "2F5496"
    write_title(
        ws1,
        f"{ws_name} - {period} Close: Reconciliation Tasks",
        f"Generated: {now} | {len(rec_tasks)} accounts | Period status: {status}",
        "2F5496", "H",
    )
    write_header_row(ws1, 4, rec_headers, "blue")
    for i, task in enumerate(rec_tasks):
        row = i + 5
        prep = user_map.get(task.get("prep_assignee", ""), task.get("prep_assignee", ""))
        rev = user_map.get(task.get("review_assignee", ""), "")
        vals = [task.get("name", ""), prep, task.get("prep_status", ""),
                task.get("prep_due", ""), rev, task.get("review_status", ""),
                task.get("review_due", ""), task.get("url", "")]
        for col, val in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            style_cell(cell, fill=status_fill(val))
    for i, w in enumerate(rec_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A5"
    ws1.auto_filter.ref = f"A4:H{4 + len(rec_tasks)}"

    # Sheet 2: Checklist Tasks
    cl_headers = ["Task Name", "Preparer", "Prep Status", "Prep Due",
                  "Reviewer", "Review Status", "Review Due"]
    cl_widths = [60, 28, 14, 12, 28, 14, 12]
    ws2 = wb.create_sheet("Checklist Tasks")
    ws2.sheet_properties.tabColor = "C55A11"
    write_title(
        ws2,
        f"{ws_name} - {period} Close: Checklist Tasks",
        f"Generated: {now} | {len(checklist_tasks)} tasks | Period status: {status}",
        "C55A11", "G",
    )
    write_header_row(ws2, 4, cl_headers, "orange")
    for i, task in enumerate(checklist_tasks):
        row = i + 5
        prep = user_map.get(task.get("prep_assignee", ""), task.get("prep_assignee", ""))
        rev = user_map.get(task.get("review_assignee", ""), "")
        vals = [task.get("name", ""), prep, task.get("prep_status", ""),
                task.get("prep_due", ""), rev, task.get("review_status", ""),
                task.get("review_due", "")]
        for col, val in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            style_cell(cell, fill=status_fill(val))
    for i, w in enumerate(cl_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A5"
    ws2.auto_filter.ref = f"A4:G{4 + len(checklist_tasks)}"

    # Sheet 3: Activity Log
    build_activity_log(wb, events, user_map, ws_name, period)

    # Sheet 4: Review Notes & Comments
    build_comments_sheet(wb, comments, user_map, ws_name, period)

    # Sheet 5: Summary
    build_summary(wb, rec_tasks, checklist_tasks, events, comments, user_map, ws_name, period)

    wb.save(args.output)
    print(json.dumps({
        "status": "success",
        "output": args.output,
        "rec_tasks": len(rec_tasks),
        "checklist_tasks": len(checklist_tasks),
        "events": len(events),
        "comments": len(comments),
        "sheets": wb.sheetnames,
    }))


if __name__ == "__main__":
    main()
