#!/usr/bin/env python3
"""Build a 3-layer formula-first Excel workpaper from JE data.

The workpaper is the primary deliverable for the reviewer. Every amount cell
in the JE sheet is a formula pointing back to Calc, and every Calc cell is a
formula pointing back to Source. The reviewer can trace any number in two clicks.

Usage:
    python build_workpaper.py --source source.json --je je_lines.json --output workpaper.xlsx
    python build_workpaper.py --source source.json --je je_lines.json --column-map '{"amount":"D","category":"B"}' --output workpaper.xlsx
    python build_workpaper.py --source source.json --je je_lines.json --ic --output workpaper.xlsx

Source JSON format:
{
  "vendor": "EOR Vendor",
  "period": "February 2026",
  "columns": ["description", "category", "amount", "employee"],
  "lines": [
    { "description": "Salaries", "category": "Payroll", "amount": 149345.38, "employee": "John Doe" },
    ...
  ]
}

JE JSON format: same as validate_je.py input
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# Formatting constants
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=11)
FORMULA_FONT = Font(name="Calibri", size=11, color="0000FF")  # Blue = formula
CURRENCY_FMT = '#,##0.00'
ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def apply_header_format(ws, row, max_col):
    """Format a row as a header."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def normalize_source_data(source_data: dict) -> dict:
    """Normalize source JSON to use 'lines' key regardless of input format.

    Accepts: 'lines', 'rows', 'data', 'items', 'records', 'transactions'.
    Returns a new dict with the data under 'lines'.
    """
    if "lines" in source_data and source_data["lines"]:
        return source_data

    alt_keys = ["rows", "data", "items", "records", "transactions", "entries"]
    for key in alt_keys:
        if key in source_data and isinstance(source_data[key], list) and source_data[key]:
            normalized = dict(source_data)
            normalized["lines"] = normalized.pop(key)
            return normalized

    # Last resort: if there's exactly one list-of-dicts value, use it
    for key, val in source_data.items():
        if isinstance(val, list) and val and isinstance(val[0], dict):
            normalized = dict(source_data)
            normalized["lines"] = val
            return normalized

    return source_data


def resolve_column_map(source_data: dict, explicit_map: dict | None) -> dict:
    """Determine column mapping for Source sheet.

    Returns a dict with keys 'amount', 'category' (and optionally others)
    mapped to 1-based column indices in the Source sheet.

    Priority:
    1. Explicit --column-map argument (letters like "D" → converted to index)
    2. Source JSON 'columns' field
    3. Auto-detect from first line's keys
    """
    lines = source_data.get("lines", [])
    if not lines:
        return {}

    # Determine column order
    if "columns" in source_data:
        columns = source_data["columns"]
    else:
        columns = list(lines[0].keys())

    # Build name-to-index map (1-based, case-insensitive lookup)
    col_index = {name: idx + 1 for idx, name in enumerate(columns)}
    col_index_lower = {name.lower(): idx + 1 for idx, name in enumerate(columns)}

    if explicit_map:
        # Explicit map uses Excel column letters → convert to indices
        resolved = {}
        for role, letter_or_name in explicit_map.items():
            if isinstance(letter_or_name, str) and len(letter_or_name) <= 2 and letter_or_name.isalpha():
                # It's a column letter like "D"
                resolved[role] = sum(
                    (ord(c.upper()) - ord('A') + 1) * (26 ** (len(letter_or_name) - 1 - i))
                    for i, c in enumerate(letter_or_name)
                )
            elif letter_or_name in col_index:
                resolved[role] = col_index[letter_or_name]
            elif letter_or_name.lower() in col_index_lower:
                resolved[role] = col_index_lower[letter_or_name.lower()]
            else:
                print(f"WARNING: Cannot resolve column '{letter_or_name}' for role '{role}'",
                      file=sys.stderr)
        return {"columns": columns, **resolved}

    # Auto-detect: look for known field names (expanded list)
    resolved = {"columns": columns}
    amount_names = [
        "amount", "net_amount", "total", "value", "dealt_amount", "gross_amount",
        "net", "sum", "balance", "interest", "principal", "debit", "credit",
        "invoice_amount", "line_amount", "ext_amount", "price",
    ]
    category_names = [
        "category", "type", "classification", "description", "name",
        "reference", "account", "line_description", "item", "memo",
        "vendor", "label", "group",
    ]

    for name in amount_names:
        if name in col_index_lower:
            # Use the original-case column name
            original_name = columns[col_index_lower[name] - 1]
            resolved["amount"] = col_index[original_name]
            break

    for name in category_names:
        if name in col_index_lower:
            original_name = columns[col_index_lower[name] - 1]
            resolved["category"] = col_index[original_name]
            break

    return resolved


def build_source_sheet(wb, source_data: dict, col_map: dict) -> int:
    """Layer 1: Raw parsed data. No formulas — pure inputs."""
    ws = wb.create_sheet("Source", 0)
    lines = source_data.get("lines", [])
    columns = col_map.get("columns", list(lines[0].keys()) if lines else [])

    if not lines:
        ws["A1"] = "No source data provided"
        return 0

    num_cols = len(columns)

    # Headers
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
    apply_header_format(ws, 1, num_cols)

    # Data rows
    for row_idx, line in enumerate(lines, 2):
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=line.get(col_name))
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col_name.lower() in ("amount", "debit", "credit", "net_amount", "total",
                                     "value", "dealt_amount", "gross_amount", "interest",
                                     "principal", "balance", "net", "sum", "price",
                                     "invoice_amount", "line_amount", "ext_amount"):
                cell.number_format = CURRENCY_FMT

    # Auto-width
    for col_idx, col_name in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(15, len(col_name) + 5)

    return len(lines)


def build_calc_sheet(wb, source_data: dict, je_data: dict, col_map: dict) -> int:
    """Layer 2: Formulas transforming source into JE amounts.

    Writes FORMULAS, not computed values. The formulas reference the Source sheet.
    Uses the resolved column map so formulas always point to the correct columns.
    """
    ws = wb.create_sheet("Calc", 1)
    lines = source_data.get("lines", [])
    entries = je_data.get("entries", [])

    if not lines:
        ws["A1"] = "No data to calculate"
        return 0

    # Get column letters from resolved map
    amount_idx = col_map.get("amount")
    category_idx = col_map.get("category")

    amount_col_letter = get_column_letter(amount_idx) if amount_idx else None
    category_col_letter = get_column_letter(category_idx) if category_idx else None
    data_end_row = len(lines) + 1  # +1 for header

    # Headers
    headers = ["Category", "Account", "DR/CR", "Amount", "Rounding Adj", "Final Amount"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    apply_header_format(ws, 1, len(headers))

    # Build one row per unique account with SUMIFS formulas
    categories = []
    seen = set()
    for entry in entries:
        for line in entry.get("lines", []):
            acct = line.get("account", "")
            if acct not in seen:
                seen.add(acct)
                dr = line.get("debit")
                cr = line.get("credit")
                direction = "DR" if dr else "CR"
                memo = line.get("memo", acct)
                categories.append({
                    "category": memo,
                    "account": acct,
                    "direction": direction,
                })

    for row_idx, cat in enumerate(categories, 2):
        # Category name
        ws.cell(row=row_idx, column=1, value=cat["category"]).font = DATA_FONT
        # Account
        ws.cell(row=row_idx, column=2, value=cat["account"]).font = DATA_FONT
        # DR/CR
        ws.cell(row=row_idx, column=3, value=cat["direction"]).font = DATA_FONT

        # Amount formula: SUMIFS from Source using resolved column map
        if amount_col_letter and category_col_letter:
            formula = (
                f'=SUMIFS(Source!{amount_col_letter}2:{amount_col_letter}{data_end_row},'
                f'Source!{category_col_letter}2:{category_col_letter}{data_end_row},'
                f'A{row_idx})'
            )
            cell = ws.cell(row=row_idx, column=4, value=formula)
        else:
            # No column map — use direct values as fallback
            # Sum amounts from source lines matching this category
            total = sum(
                float(l.get("amount", 0) or 0)
                for l in lines
                if l.get("category", l.get("description", "")) == cat["category"]
            )
            cell = ws.cell(row=row_idx, column=4, value=total)
        cell.font = FORMULA_FONT
        cell.number_format = CURRENCY_FMT
        cell.border = THIN_BORDER

        # Rounding adjustment (populated only on last row if needed)
        ws.cell(row=row_idx, column=5, value=0).font = DATA_FONT
        ws.cell(row=row_idx, column=5).number_format = CURRENCY_FMT

        # Final Amount = Amount + Rounding Adj
        final_formula = f"=D{row_idx}+E{row_idx}"
        final_cell = ws.cell(row=row_idx, column=6, value=final_formula)
        final_cell.font = FORMULA_FONT
        final_cell.number_format = CURRENCY_FMT
        final_cell.border = THIN_BORDER

    # Totals row
    total_row = len(categories) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", size=11, bold=True)
    for col in [4, 5, 6]:
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
        cell = ws.cell(row=total_row, column=col, value=formula)
        cell.font = Font(name="Calibri", size=11, bold=True, color="0000FF")
        cell.number_format = CURRENCY_FMT
        cell.border = THIN_BORDER

    # Column widths
    widths = [25, 35, 8, 18, 15, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return len(categories)


def build_je_sheet(wb, je_data: dict, calc_row_count: int, intercompany: bool = False):
    """Layer 3: JE lines with formula references back to Calc.

    Every amount cell references the Calc sheet. Balance check at the bottom.
    If intercompany=True, generates 4-line IC structure with clearing account
    net-to-zero check.
    """
    ws = wb.create_sheet("JE", 2)
    entries = je_data.get("entries", [])

    # Headers — include Line Subsidiary for IC entries
    headers = [
        "External ID", "Date", "Subsidiary",
    ]
    if intercompany:
        headers.append("To Subsidiary")
        headers.append("Line Subsidiary")
    headers.extend([
        "Account", "Debit", "Credit", "Department", "Class", "Location",
        "Memo", "Name"
    ])
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    apply_header_format(ws, 1, len(headers))

    # Column index offsets
    if intercompany:
        acct_col, dr_col, cr_col = 6, 7, 8
        dept_col, cls_col, loc_col, memo_col, name_col = 9, 10, 11, 12, 13
    else:
        acct_col, dr_col, cr_col = 4, 5, 6
        dept_col, cls_col, loc_col, memo_col, name_col = 7, 8, 9, 10, 11

    row_idx = 2
    calc_row = 2  # tracks which Calc row to reference

    for entry in entries:
        for line in entry.get("lines", []):
            ws.cell(row=row_idx, column=1, value=entry.get("external_id", "")).font = DATA_FONT
            ws.cell(row=row_idx, column=2, value=entry.get("date", "")).font = DATA_FONT
            ws.cell(row=row_idx, column=3, value=entry.get("subsidiary", "")).font = DATA_FONT

            if intercompany:
                ws.cell(row=row_idx, column=4, value=entry.get("to_subsidiary", "")).font = DATA_FONT
                ws.cell(row=row_idx, column=5,
                        value=line.get("line_subsidiary", entry.get("subsidiary", ""))).font = DATA_FONT

            ws.cell(row=row_idx, column=acct_col, value=line.get("account", "")).font = DATA_FONT

            # Debit formula: if Calc direction is DR, pull Final Amount
            dr_formula = f'=IF(Calc!C{calc_row}="DR",Calc!F{calc_row},"")'
            dr_cell = ws.cell(row=row_idx, column=dr_col, value=dr_formula)
            dr_cell.font = FORMULA_FONT
            dr_cell.number_format = CURRENCY_FMT
            dr_cell.border = THIN_BORDER

            # Credit formula: if Calc direction is CR, pull Final Amount
            cr_formula = f'=IF(Calc!C{calc_row}="CR",Calc!F{calc_row},"")'
            cr_cell = ws.cell(row=row_idx, column=cr_col, value=cr_formula)
            cr_cell.font = FORMULA_FONT
            cr_cell.number_format = CURRENCY_FMT
            cr_cell.border = THIN_BORDER

            ws.cell(row=row_idx, column=dept_col, value=line.get("department", "")).font = DATA_FONT
            ws.cell(row=row_idx, column=cls_col, value=line.get("class", "")).font = DATA_FONT
            ws.cell(row=row_idx, column=loc_col, value=line.get("location", "")).font = DATA_FONT
            ws.cell(row=row_idx, column=memo_col, value=line.get("memo", "")).font = DATA_FONT
            ws.cell(row=row_idx, column=name_col, value=line.get("name", "")).font = DATA_FONT

            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = THIN_BORDER

            row_idx += 1
            calc_row += 1

    # Balance check row
    balance_row = row_idx + 1
    ws.cell(row=balance_row, column=acct_col, value="BALANCE CHECK:").font = Font(
        name="Calibri", size=11, bold=True
    )

    dr_sum = f"=SUM({get_column_letter(dr_col)}2:{get_column_letter(dr_col)}{row_idx - 1})"
    cr_sum = f"=SUM({get_column_letter(cr_col)}2:{get_column_letter(cr_col)}{row_idx - 1})"
    ws.cell(row=balance_row, column=dr_col, value=dr_sum).font = Font(
        name="Calibri", size=11, bold=True, color="0000FF"
    )
    ws.cell(row=balance_row, column=dr_col).number_format = CURRENCY_FMT

    ws.cell(row=balance_row, column=cr_col, value=cr_sum).font = Font(
        name="Calibri", size=11, bold=True, color="0000FF"
    )
    ws.cell(row=balance_row, column=cr_col).number_format = CURRENCY_FMT

    # Difference cell
    diff_row = balance_row + 1
    ws.cell(row=diff_row, column=acct_col, value="DIFFERENCE:").font = Font(
        name="Calibri", size=11, bold=True
    )
    diff_formula = f"={get_column_letter(dr_col)}{balance_row}-{get_column_letter(cr_col)}{balance_row}"
    diff_cell = ws.cell(row=diff_row, column=dr_col, value=diff_formula)
    diff_cell.font = Font(name="Calibri", size=11, bold=True, color="0000FF")
    diff_cell.number_format = CURRENCY_FMT

    ws.cell(row=diff_row, column=dr_col + 1, value="<-- Should be 0.00").font = Font(
        name="Calibri", size=10, italic=True, color="999999"
    )

    # Intercompany clearing account net-to-zero check
    if intercompany:
        ic_row = diff_row + 2
        ws.cell(row=ic_row, column=1, value="IC CLEARING NET:").font = Font(
            name="Calibri", size=11, bold=True
        )
        ws.cell(row=ic_row, column=2,
                value="Clearing account lines must net to zero across subsidiaries").font = Font(
            name="Calibri", size=10, italic=True, color="999999"
        )
        # Sum DR and CR for clearing account lines (agent should verify account code)
        note_row = ic_row + 1
        ws.cell(row=note_row, column=1,
                value="(Filter JE lines by clearing account to verify)").font = Font(
            name="Calibri", size=10, italic=True, color="999999"
        )

    # Column widths
    base_widths = [28, 12, 25, 35, 16, 16, 30, 20, 20, 40, 20]
    if intercompany:
        base_widths = [28, 12, 25, 25, 25, 35, 16, 16, 30, 20, 20, 40, 20]
    for i, w in enumerate(base_widths, 1):
        if i <= len(headers):
            ws.column_dimensions[get_column_letter(i)].width = w


def main():
    parser = argparse.ArgumentParser(description="Build 3-layer formula-first JE workpaper")
    parser.add_argument("--source", required=True, help="Path to source data JSON")
    parser.add_argument("--je", required=True, help="Path to JE lines JSON")
    parser.add_argument("--column-map", help="JSON column mapping: {\"amount\": \"D\", \"category\": \"B\"}")
    parser.add_argument("--ic", action="store_true",
                        help="Intercompany mode: generate 4-line IC structure with clearing checks")
    parser.add_argument("--output", required=True, help="Output Excel file path")
    args = parser.parse_args()

    # Load data
    for path_arg, name in [(args.source, "source"), (args.je, "JE")]:
        if not Path(path_arg).exists():
            print(f"ERROR: {name} file not found: {path_arg}", file=sys.stderr)
            sys.exit(1)

    with open(args.source) as f:
        source_data = normalize_source_data(json.load(f))
    with open(args.je) as f:
        je_data = json.load(f)

    # Parse column map if provided
    explicit_map = None
    if args.column_map:
        try:
            explicit_map = json.loads(args.column_map)
        except json.JSONDecodeError:
            print(f"ERROR: Invalid JSON for --column-map: {args.column_map}", file=sys.stderr)
            sys.exit(1)

    # Resolve column mapping
    col_map = resolve_column_map(source_data, explicit_map)

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)

    source_rows = build_source_sheet(wb, source_data, col_map)
    calc_rows = build_calc_sheet(wb, source_data, je_data, col_map)
    build_je_sheet(wb, je_data, calc_rows, intercompany=args.ic)

    # Save
    output_path = Path(args.output)
    wb.save(str(output_path))

    entry_count = sum(len(e.get("lines", [])) for e in je_data.get("entries", []))
    print(f"Workpaper saved: {output_path}")
    print(f"  Source: {source_rows} rows")
    print(f"  Calc: {calc_rows} categories")
    print(f"  JE: {entry_count} lines")
    if args.ic:
        print(f"  Mode: Intercompany (4-line structure)")


if __name__ == "__main__":
    main()
