"""
rec-to-numeric — convert a Balance Sheet Rec assignments file (xlsx or csv)
into a Numeric Rec Assignment import xlsx.

Output columns (Numeric's Recon assignment template):
  Account number, Subsidiary, Internal ID of Entity, Category, Property,
  Preparer Email, Due Day of Close Preparer,
  Reviewer Email, Due Day of Close Reviewer,
  Second Reviewer Email, Due Day of Close Second Reviewer

The script does five things:
  1. Read assignments  (xlsx or csv, auto-detects the header row)
  2. Look up Internal ID of Entity from an ids file
  3. Match each row's account # against the current Numeric COA
  4. Map preparer/reviewer names to emails
  5. Apply Numeric's import rules and write a 3-tab output xlsx
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from collections import Counter

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# 1. Header detection
# ---------------------------------------------------------------------------

SYNONYMS = {
    "entity":      ["entity", "subsidiary", "company", "legal entity"],
    "account":     ["gl account", "account number", "account #", "account", "acct"],
    "description": ["description", "account description", "desc"],
    "team":        ["team", "responsible team"],
    "preparer":    ["preparer", "preparer name", "owner", "prepared by"],
    "reviewer":    ["reviewer", "reviewer name", "approver", "reviewed by"],
    "notes":       ["notes", "comments", "remarks"],
}


def _read_assignments(path: Path, sheet: str | None):
    """Return (rows: list[list], sheet_label: str). Each row is a list of
    cell values matching the header row's width. Handles xlsx + csv."""
    if path.suffix.lower() == ".csv":
        with open(path, encoding="utf-8-sig", newline="") as f:
            return [row for row in csv.reader(f)], "csv"
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows, ws.title


def find_header_row(rows: list[list], must_have=("entity", "account")):
    """Return ({field: column_idx_0based}, header_row_idx)."""
    syn_flat = {syn.lower(): k for k, vals in SYNONYMS.items() for syn in vals}
    for r_idx, row in enumerate(rows[:25]):
        hits = {}
        for c_idx, v in enumerate(row):
            if v is None:
                continue
            key = str(v).strip().lower()
            if key in syn_flat:
                hits[syn_flat[key]] = c_idx
        if all(f in hits for f in must_have):
            return hits, r_idx
    raise RuntimeError(
        "Couldn't find a header row containing entity + account in the "
        "first 25 rows. Run with --list-headers to see what's in the file."
    )


# ---------------------------------------------------------------------------
# 2. Entity IDs
# ---------------------------------------------------------------------------

def load_entity_ids(path: Path) -> dict[str, object]:
    """Return entity_name → internal_id. Supports two shapes:
      A) clean 2-col table with 'Entity' and 'ID' headers
      B) single-column with 5-row-per-entity blocks
         (entity, acronym, currency, accounts, id)
    """
    if path.suffix.lower() == ".csv":
        with open(path, encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            out = {}
            for row in reader:
                e = row.get("Entity") or row.get("entity")
                i = row.get("ID") or row.get("id")
                if e and i:
                    out[e.strip()] = i
            return out

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Shape A — header row with 'entity' + 'id'
    for r in range(1, min(ws.max_row, 10) + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        lowered = [str(v).strip().lower() if v is not None else "" for v in row]
        if "entity" in lowered and "id" in lowered:
            e_col = lowered.index("entity") + 1
            i_col = lowered.index("id") + 1
            out = {}
            for rr in range(r + 1, ws.max_row + 1):
                e = ws.cell(rr, e_col).value
                i = ws.cell(rr, i_col).value
                if e and i is not None:
                    out[str(e).strip()] = i
            if out:
                return out

    # Shape B — 5-row blocks in column A
    vals = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
    out = {}
    i = 0
    while i + 4 < len(vals):
        block = vals[i:i+5]
        entity, _, _, _, eid = block
        if entity and eid is not None:
            out[str(entity).strip()] = eid
        i += 5
    return out


# ---------------------------------------------------------------------------
# 3. COA + alias CSV
# ---------------------------------------------------------------------------

def load_coa(path: Path | None) -> tuple[dict[str, str], set[str]]:
    """Return (description_lower → code, set_of_valid_codes)."""
    if not path:
        return {}, set()
    name_to_code, valid_codes = {}, set()
    with open(path, encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            code = (row.get("code") or "").strip()
            name = (row.get("name") or "").strip()
            if code:
                valid_codes.add(code)
            if code and name:
                name_to_code.setdefault(name.lower(), code)
    return name_to_code, valid_codes


def load_aliases(path: Path | None) -> dict[str, str]:
    """Load a 2-column 'old_desc, new_desc' CSV (header optional).
    Both sides lowercased."""
    if not path:
        return {}
    out = {}
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) < 2:
                continue
            old, new = row[0].strip(), row[1].strip()
            # Skip a likely header row
            if old.lower() in ("old_desc", "old description", "old"):
                continue
            if old and new:
                out[old.lower()] = new.lower()
    return out


def match_account(desc: str | None, coa: dict[str, str],
                  aliases: dict[str, str]) -> str | None:
    """Try exact match → alias → prefix match. Return new account code or None."""
    if not desc:
        return None
    key = desc.lower().strip()

    # Exact
    if key in coa:
        return coa[key]

    # Alias rewrite, then exact
    if key in aliases and aliases[key] in coa:
        return coa[aliases[key]]

    # Prefix (handles BS-rec descriptions truncated at fixed column width)
    if len(key) >= 20:
        candidates = [name for name in coa if name.startswith(key)]
        if candidates:
            return coa[min(candidates, key=len)]

    return None


# ---------------------------------------------------------------------------
# 4. Names → emails
# ---------------------------------------------------------------------------

NON_PERSON = {"n/a", "na", "all", "fp&a", "pinky"}


def load_email_map(path: Path | None) -> dict[str, str]:
    """name,email CSV (header optional). Lowercase keys + values."""
    if not path:
        return {}
    out = {}
    with open(path, encoding="utf-8-sig", newline="") as f:
        for row in csv.reader(f):
            if len(row) < 2:
                continue
            name, email = row[0].strip(), row[1].strip()
            if "@" not in email:
                continue
            out[name.lower()] = email.lower()
    return out


def split_names(cell) -> list[str]:
    if cell is None:
        return []
    s = str(cell).strip()
    if not s:
        return []
    s = re.sub(r"\bN/A\b", "__NA__", s, flags=re.IGNORECASE)
    parts = re.split(r"\s*/\s*|\s*;\s*|\s*,\s*|\s+and\s+", s)
    out = []
    for p in parts:
        p = p.replace("__NA__", "N/A").strip()
        p = re.sub(r"\s*\(.*?\).*$", "", p).strip()
        if p:
            out.append(p)
    return out


def resolve_email(name: str | None, email_map: dict[str, str]
                  ) -> tuple[str, str | None]:
    """Return (email, issue_or_None)."""
    if not name:
        return "", None
    k = name.lower().strip()
    if k in ("n/a", "na"):
        return "", None
    if k in NON_PERSON:
        return "", f"non-person value '{name}'"
    k = re.sub(r"\s*\(.*?\)", "", k).strip()
    if k in email_map:
        return email_map[k], None
    return "", f"no email mapping for '{name}'"


# ---------------------------------------------------------------------------
# 5. Main
# ---------------------------------------------------------------------------

OUT_COLS = [
    "Account number", "Subsidiary", "Internal ID of Entity",
    "Category", "Property",
    "Preparer Email", "Due Day of Close Preparer",
    "Reviewer Email", "Due Day of Close Reviewer",
    "Second Reviewer Email", "Due Day of Close Second Reviewer",
]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--assignments", required=True, type=Path,
                    help="Source xlsx or csv with rec assignments")
    ap.add_argument("--assignments-sheet", default=None)
    ap.add_argument("--ids", type=Path,
                    help="Entity-ID lookup file (xlsx or csv). Required "
                         "unless --list-headers.")
    ap.add_argument("--coa", type=Path,
                    help="Numeric COA CSV (columns: code, name, ...). "
                         "Pull via list_financial_accounts.")
    ap.add_argument("--aliases", type=Path,
                    help="Optional 2-column CSV: 'old description, new description'. "
                         "Use when BS Rec descriptions don't match the COA.")
    ap.add_argument("--email-map", type=Path,
                    help="Optional 2-column CSV: 'name, email'.")
    ap.add_argument("--property", default="Team",
                    help="Source column header to feed into Property "
                         "(default: Team)")
    ap.add_argument("--prep-due", type=int, default=None)
    ap.add_argument("--rev-due", type=int, default=None)
    ap.add_argument("--rev2-due", type=int, default=None,
                    help="Falls back to --rev-due if omitted")
    ap.add_argument("--out", type=Path,
                    help="Output xlsx path. Required unless --list-headers.")
    ap.add_argument("--list-headers", action="store_true",
                    help="Print the source file's first 10 rows and the "
                         "auto-detected column mapping, then exit.")
    args = ap.parse_args()

    rows, sheet_label = _read_assignments(args.assignments,
                                          args.assignments_sheet)

    if args.list_headers:
        print(f"Sheet: {sheet_label}  ({len(rows)} rows)\n")
        print("First 10 rows:")
        for i, r in enumerate(rows[:10]):
            print(f"  R{i}: {r}")
        try:
            cols, hdr_row = find_header_row(rows)
            print(f"\nDetected header row: {hdr_row}")
            print(f"Column mapping: {cols}")
        except RuntimeError as e:
            print(f"\n⚠ {e}")
        return

    if not args.ids or not args.out:
        ap.error("--ids and --out are required (omit only with --list-headers)")

    cols, hdr_row = find_header_row(rows)
    print(f"Sheet: {sheet_label}  | header row: {hdr_row}")
    print(f"Column mapping: {cols}")

    ids = load_entity_ids(args.ids)
    coa, valid_codes = load_coa(args.coa)
    aliases = load_aliases(args.aliases)
    email_map = load_email_map(args.email_map)
    if args.coa:
        print(f"COA: {len(coa)} accounts, {len(valid_codes)} valid codes")
    if aliases:
        print(f"Aliases: {len(aliases)} renames loaded")

    # Find the "Property" column by header label
    prop_col = None
    if hdr_row < len(rows):
        for c, v in enumerate(rows[hdr_row]):
            if v and str(v).strip().lower() == args.property.strip().lower():
                prop_col = c
                break

    out_rows = []
    swapped = unmapped_count = multi_drop = 0

    for r_idx in range(hdr_row + 1, len(rows)):
        row = rows[r_idx]

        def col(field):
            c = cols.get(field)
            return row[c] if c is not None and c < len(row) else None

        entity = col("entity")
        acct = col("account")
        desc = col("description") or ""
        prop_val = (row[prop_col] if prop_col is not None and prop_col < len(row) else "")
        prep_c = col("preparer")
        rev_c = col("reviewer")

        if not any([entity, acct, desc, prop_val, prep_c, rev_c]):
            continue

        entity = str(entity).strip() if entity else ""
        notes = []
        unmapped = False

        # Account renumbering
        final_acct = acct
        if coa:
            new_code = match_account(desc, coa, aliases)
            if new_code:
                if str(new_code) != str(acct):
                    swapped += 1
                final_acct = new_code
            elif valid_codes and str(acct).strip() in valid_codes:
                pass  # existing code is already valid
            else:
                unmapped = True
                unmapped_count += 1
                notes.append(f"no match in current COA for '{desc}' — kept old #{acct}")

        # Entity ID
        eid = ids.get(entity, "")
        if not eid and entity:
            notes.append(f"entity '{entity}' not in ids file")

        # Preparer (one slot)
        prep_list = split_names(prep_c)
        p_email = ""
        if prep_list:
            p_email, p_iss = resolve_email(prep_list[0], email_map)
            if p_iss:
                notes.append(f"preparer: {p_iss}")
            if len(prep_list) > 1:
                multi_drop += 1
                notes.append(
                    f"additional preparer(s) dropped: {', '.join(prep_list[1:])}")

        # Reviewer (1 + optional 2nd)
        rev_list = split_names(rev_c)
        r_email = r2_email = ""
        if rev_list:
            r_email, r_iss = resolve_email(rev_list[0], email_map)
            if r_iss:
                notes.append(f"reviewer: {r_iss}")
            if len(rev_list) > 1:
                r2_email, r2_iss = resolve_email(rev_list[1], email_map)
                if r2_iss:
                    notes.append(f"second reviewer: {r2_iss}")
                if len(rev_list) > 2:
                    multi_drop += 1
                    notes.append(
                        f"3rd+ reviewer(s) dropped: {', '.join(rev_list[2:])}")

        # No preparer → no reviewer chain
        if not p_email:
            r_email = r2_email = ""

        # Due days only when matching email is present
        prep_due = args.prep_due if (args.prep_due is not None and p_email) else ""
        rev_due  = args.rev_due  if (args.rev_due  is not None and r_email) else ""
        r2_default = args.rev2_due if args.rev2_due is not None else args.rev_due
        rev2_due = r2_default if (r2_default is not None and r2_email) else ""

        out_rows.append({
            "Account number": final_acct,
            "Subsidiary": entity,
            "Internal ID of Entity": eid,
            "Category": args.property,
            "Property": prop_val or "",
            "Preparer Email": p_email,
            "Due Day of Close Preparer": prep_due,
            "Reviewer Email": r_email,
            "Due Day of Close Reviewer": rev_due,
            "Second Reviewer Email": r2_email,
            "Due Day of Close Second Reviewer": rev2_due,
            "_row": r_idx + 1,
            "_notes": notes,
            "_desc": desc,
            "_unmapped": unmapped,
        })

    write_xlsx(out_rows, args.out)

    importable = sum(1 for r in out_rows if not r["_unmapped"])
    flagged = sum(1 for r in out_rows
                  if r["_notes"] and not r["_unmapped"])
    per_sub = Counter(r["Subsidiary"] for r in out_rows if not r["_unmapped"])

    print(f"\n✓ {len(out_rows)} source rows processed")
    print(f"  → {importable} rows on 'Import' tab (ready for Numeric)")
    print(f"  → {unmapped_count} rows on 'Unmapped to Numeric' tab")
    print(f"  • {flagged} importable rows flagged in Notes & Warnings")
    print(f"  • {swapped} account numbers swapped (matched to current COA)")
    print(f"  • {multi_drop} cells had extra assignees beyond what Numeric supports")
    print(f"\n  Rows per subsidiary (importable):")
    for e, n in per_sub.most_common():
        print(f"     {n:>3}  {e}  (id={ids.get(e, '?')})")
    print(f"\n  Output: {args.out}")


# ---------------------------------------------------------------------------
# Output writer
# ---------------------------------------------------------------------------

def write_xlsx(rows: list[dict], out_path: Path):
    importable = [r for r in rows if not r["_unmapped"]]
    unmapped   = [r for r in rows if r["_unmapped"]]

    wb = Workbook()
    hfont = Font(bold=True, color="FFFFFF")
    fill_main = PatternFill("solid", fgColor="2F4858")
    fill_warn = PatternFill("solid", fgColor="B85450")
    fill_unmap = PatternFill("solid", fgColor="C97B4F")

    # Tab 1: Import
    ws = wb.active
    ws.title = "Import"
    for c, name in enumerate(OUT_COLS, 1):
        cell = ws.cell(1, c, name)
        cell.font = hfont; cell.fill = fill_main
        cell.alignment = Alignment(horizontal="left")
    for i, r in enumerate(importable, 2):
        for c, name in enumerate(OUT_COLS, 1):
            ws.cell(i, c, r[name])

    # Tab 2: Notes & Warnings
    ws2 = wb.create_sheet("Notes & Warnings")
    hdr = ["Source Row", "Subsidiary", "Account #", "Description", "Notes"]
    for c, name in enumerate(hdr, 1):
        cell = ws2.cell(1, c, name)
        cell.font = hfont; cell.fill = fill_warn
        cell.alignment = Alignment(horizontal="left")
    i = 2
    for r in importable:
        if not r["_notes"]:
            continue
        ws2.cell(i, 1, r["_row"])
        ws2.cell(i, 2, r["Subsidiary"])
        ws2.cell(i, 3, r["Account number"])
        ws2.cell(i, 4, r["_desc"])
        ws2.cell(i, 5, " | ".join(r["_notes"]))
        i += 1

    # Tab 3: Unmapped to Numeric
    ws3 = wb.create_sheet("Unmapped to Numeric")
    hdr3 = ["Source Row", "Subsidiary", "Old Account #", "Old Description",
            "Property", "Preparer (raw)", "Reviewer (raw)", "Notes"]
    for c, name in enumerate(hdr3, 1):
        cell = ws3.cell(1, c, name)
        cell.font = hfont; cell.fill = fill_unmap
        cell.alignment = Alignment(horizontal="left")
    for i, r in enumerate(unmapped, 2):
        ws3.cell(i, 1, r["_row"])
        ws3.cell(i, 2, r["Subsidiary"])
        ws3.cell(i, 3, r["Account number"])
        ws3.cell(i, 4, r["_desc"])
        ws3.cell(i, 5, r["Property"])
        ws3.cell(i, 6, r["Preparer Email"])
        ws3.cell(i, 7, r["Reviewer Email"])
        ws3.cell(i, 8, " | ".join(r["_notes"]))

    for w_idx, widths in enumerate([
        [14, 12, 10, 14, 30, 32, 8, 32, 8, 32, 8],
        [8, 12, 12, 45, 55],
        [8, 12, 12, 45, 18, 32, 32, 55],
    ]):
        sheet = [ws, ws2, ws3][w_idx]
        for i, w in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = w
        sheet.freeze_panes = "A2"

    wb.save(out_path)


if __name__ == "__main__":
    main()
