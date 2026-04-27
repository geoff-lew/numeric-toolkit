#!/usr/bin/env python3
"""
generate_outputs.py — Generates the accrual workpaper (xlsx), JE CSV, and validation file.

Usage:
    python3 generate_outputs.py \
        --spend-matrix <vendor_spend_matrix.json> \
        --confirmed-vendors <confirmed_vendors.json> \
        --output-dir <dir> \
        --period-slug <mar-2025> \
        --period-end <3/31/2025> \
        --expense-acct-code <5635> \
        --expense-acct-name "Office Expense : Software Expense" \
        --accrued-exp-acct-code <2300> \
        --entity-id <1>

confirmed_vendors.json format:
    [
        {
            "vendor": "Vendor Name",
            "method": "6-mo avg",
            "method_detail": "6-month avg of (Sep $X + Oct $Y + ...)",
            "proposed_amount": 1234.56,
            "trigger": "$0 in Mar, $5,000 in Feb"
        }
    ]

Outputs:
    - {slug}_accrual_workpaper_{period}.xlsx  (3-sheet workpaper)
    - {slug}_journal_entries_{period}.csv     (NetSuite-ready JE import)
    - {slug}_accrual_validation_{period}.txt  (validation + balance check)
"""

import json
import csv
import argparse
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl --break-system-packages")
    sys.exit(1)


def build_workpaper(
    vendors_data: dict,
    confirmed: list,
    completed_months: list,
    acct_name: str,
    accrued_exp_code: str,
    entity_id: str,
    period_end: str,
    period_slug: str,
    output_path: str,
):
    """Build the 3-sheet Excel workpaper."""
    wb = Workbook()
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold = Font(bold=True)
    blue = Font(color="0000FF")
    money = "#,##0.00"

    confirmed_names = {c["vendor"] for c in confirmed}
    confirmed_lookup = {c["vendor"]: c for c in confirmed}

    # --- Sheet 1: All vendors ---
    ws1 = wb.active
    ws1.title = f"{acct_name.split(':')[-1].strip()} by Vendor"
    ws1.sheet_properties.showGridLines = False

    ws1.cell(row=1, column=1, value=f"{acct_name} — {period_slug.title()} Accrual Workpaper").font = Font(bold=True, size=14)

    headers = ["Vendor"] + completed_months + ["6-Month Avg", "Accrual Estimate", "Accrual Memo"]
    for ci, h in enumerate(headers, 1):
        ws1.cell(row=3, column=ci, value=h).font = bold

    all_vendors = sorted(vendors_data.keys())
    row = 4
    for v in all_vendors:
        spend = vendors_data[v]
        ws1.cell(row=row, column=1, value=v)

        for mi, m in enumerate(completed_months, 2):
            ws1.cell(row=row, column=mi, value=spend.get(m, 0)).number_format = money

        vals = [spend.get(m, 0) for m in completed_months]
        avg = sum(vals) / len(vals) if vals else 0
        ws1.cell(row=row, column=len(completed_months) + 2, value=avg).number_format = money

        if v in confirmed_names:
            cv = confirmed_lookup[v]
            ws1.cell(row=row, column=len(completed_months) + 3, value=cv["proposed_amount"]).number_format = money
            ws1.cell(row=row, column=len(completed_months) + 4, value=cv.get("memo", f"{v} | {period_slug} accrual: {cv['method_detail']} = ${cv['proposed_amount']:,.2f}"))
            for ci in range(1, len(headers) + 1):
                ws1.cell(row=row, column=ci).fill = yellow
        else:
            ws1.cell(row=row, column=len(completed_months) + 3, value=0).number_format = money

        row += 1

    # Total row
    total_row = row
    ws1.cell(row=total_row, column=1, value="TOTAL").font = bold
    for ci in range(2, len(headers) + 1):
        cl = get_column_letter(ci)
        ws1.cell(row=total_row, column=ci, value=f"=SUM({cl}4:{cl}{total_row - 1})").number_format = money
        ws1.cell(row=total_row, column=ci).font = bold

    ws1.freeze_panes = "B4"
    for ci in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(ci)].width = 18 if ci > 1 else 30

    # --- Sheet 2: Accrual Detail ---
    ws2 = wb.create_sheet("Accrual Detail")
    ws2.sheet_properties.showGridLines = False

    detail_headers = ["Vendor"] + completed_months + ["Avg", "Accrual Memo"]
    for ci, h in enumerate(detail_headers, 1):
        ws2.cell(row=1, column=ci, value=h).font = bold

    row = 2
    for cv in confirmed:
        v = cv["vendor"]
        spend = vendors_data.get(v, {})
        ws2.cell(row=row, column=1, value=v)
        for mi, m in enumerate(completed_months, 2):
            ws2.cell(row=row, column=mi, value=spend.get(m, 0)).number_format = money
        avg_col = len(completed_months) + 2
        ws2.cell(row=row, column=avg_col, value=cv["proposed_amount"]).number_format = money
        ws2.cell(row=row, column=avg_col + 1, value=cv.get("memo", f"{v} | {period_slug} accrual: {cv['method_detail']} = ${cv['proposed_amount']:,.2f}"))
        row += 1

    total_s2 = row
    ws2.cell(row=total_s2, column=1, value="TOTAL").font = bold
    avg_col = len(completed_months) + 2
    avg_letter = get_column_letter(avg_col)
    ws2.cell(row=total_s2, column=avg_col, value=f"=SUM({avg_letter}2:{avg_letter}{total_s2 - 1})").number_format = money

    for ci in range(1, len(detail_headers) + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 18 if ci > 1 else 30

    # --- Sheet 3: Numeric ---
    ws3 = wb.create_sheet("Numeric")
    ws3.cell(row=1, column=1, value="Account").font = bold
    ws3.cell(row=1, column=2, value="Account Name").font = bold
    ws3.cell(row=1, column=3, value="Entity").font = bold
    ws3.cell(row=1, column=4, value=period_end).font = bold

    ws3.cell(row=2, column=1, value=int(accrued_exp_code))
    ws3.cell(row=2, column=2, value="Accrued Expenses")
    ws3.cell(row=2, column=3, value=entity_id)
    ws3.cell(row=2, column=4, value=f"='Accrual Detail'!{avg_letter}{total_s2}").number_format = money
    ws3.cell(row=2, column=4).font = blue

    for ci in range(1, 5):
        ws3.column_dimensions[get_column_letter(ci)].width = 20

    wb.save(output_path)
    return total_s2, avg_letter  # For cross-referencing in validation


def build_je_csv(
    confirmed: list,
    expense_acct_code: str,
    expense_acct_name: str,
    accrued_exp_code: str,
    period_slug: str,
    period_end: str,
    output_path: str,
) -> tuple[float, float]:
    """Build the NetSuite-ready JE CSV. Returns (total_debits, total_credits)."""
    period_label = period_slug.replace("-", "").upper()  # e.g. MAR2025

    rows = []
    for i, cv in enumerate(confirmed, 1):
        memo = f"{cv['vendor']} | {period_slug.title()} accrual: {cv['method']} = ${cv['proposed_amount']:,.2f} | Trigger: {cv['trigger']}"
        rows.append({
            "External ID": f"JE-ACCRUAL-{period_label}-{i:03d}",
            "Date": period_end,
            "Currency": "USD",
            "Account Number": expense_acct_code,
            "Account Name": expense_acct_name,
            "Debit": f"{cv['proposed_amount']:.2f}",
            "Credit": "",
            "Memo": memo,
        })
        rows.append({
            "External ID": f"JE-ACCRUAL-{period_label}-{i:03d}",
            "Date": period_end,
            "Currency": "USD",
            "Account Number": accrued_exp_code,
            "Account Name": "Accrued Expenses",
            "Debit": "",
            "Credit": f"{cv['proposed_amount']:.2f}",
            "Memo": memo,
        })

    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["External ID", "Date", "Currency", "Account Number", "Account Name", "Debit", "Credit", "Memo"])
        writer.writeheader()
        writer.writerows(rows)

    total_debits = sum(float(r["Debit"]) for r in rows if r["Debit"])
    total_credits = sum(float(r["Credit"]) for r in rows if r["Credit"])
    return total_debits, total_credits


def build_validation(
    confirmed: list,
    excluded: list,
    completed_months: list,
    vendors_data: dict,
    total_debits: float,
    total_credits: float,
    acct_name: str,
    period_slug: str,
    output_path: str,
):
    """Build the validation text file."""
    lines = [
        f"{acct_name.upper()} ACCRUAL VALIDATION — {period_slug.title()}",
        "=" * 60,
        "",
        "Section A — Accrual Candidates",
        "-" * 40,
    ]

    for cv in confirmed:
        spend = vendors_data.get(cv["vendor"], {})
        hist = " | ".join(f"{m}: ${spend.get(m, 0):,.2f}" for m in completed_months)
        lines.append(f"  {cv['vendor']}: ${cv['proposed_amount']:,.2f} ({cv['method']}) — {cv['trigger']}")
        lines.append(f"    History: {hist}")

    lines.append("")
    total = sum(c["proposed_amount"] for c in confirmed)
    lines.append(f"Total accrual: ${total:,.2f}")
    lines.append("")

    if excluded:
        lines.append("Excluded vendors:")
        for ev in excluded:
            lines.append(f"  - {ev['vendor']}: ${ev.get('proposed_amount', 0):,.2f} — {ev.get('reason', 'User chose not to accrue')}")
        lines.append("")

    lines.extend([
        "Journal Entry Balance Check",
        "=" * 40,
        f"Total debits:  ${total_debits:,.2f}",
        f"Total credits: ${total_credits:,.2f}",
        f"Balance: {'PASS' if abs(total_debits - total_credits) < 0.01 else 'FAIL'}",
        f"Workpaper total matches CSV: {'PASS' if abs(total - total_debits) < 0.01 else 'FAIL'}",
    ])

    with open(output_path, "w") as f:
        f.write("\n".join(lines))


def main():
    parser = argparse.ArgumentParser(description="Generate accrual workpaper, JE CSV, and validation")
    parser.add_argument("--spend-matrix", required=True, help="Path to vendor_spend_matrix.json")
    parser.add_argument("--confirmed-vendors", required=True, help="Path to confirmed_vendors.json")
    parser.add_argument("--output-dir", required=True, help="Output directory")
    parser.add_argument("--period-slug", required=True, help="Period slug (e.g. mar-2025)")
    parser.add_argument("--period-end", required=True, help="Period end date (e.g. 3/31/2025)")
    parser.add_argument("--open-period", required=True, help="Open period YYYY-MM (e.g. 2025-03)")
    parser.add_argument("--expense-acct-code", required=True)
    parser.add_argument("--expense-acct-name", required=True)
    parser.add_argument("--accrued-exp-acct-code", required=True, help="Accrued expenses account code (e.g. 2300)")
    parser.add_argument("--entity-id", default="", help="Entity internal ID")
    parser.add_argument("--excluded-vendors", default=None, help="Path to excluded_vendors.json (optional)")
    args = parser.parse_args()

    with open(args.spend_matrix) as f:
        matrix = json.load(f)
    with open(args.confirmed_vendors) as f:
        confirmed = json.load(f)

    excluded = []
    if args.excluded_vendors:
        with open(args.excluded_vendors) as f:
            excluded = json.load(f)

    completed_months = [m for m in matrix["months"] if m < args.open_period]
    slug = args.expense_acct_name.split(":")[-1].strip().lower().replace(" ", "_")

    os.makedirs(args.output_dir, exist_ok=True)

    wp_path = os.path.join(args.output_dir, f"{slug}_accrual_workpaper_{args.period_slug}.xlsx")
    je_path = os.path.join(args.output_dir, f"{slug}_journal_entries_{args.period_slug}.csv")
    val_path = os.path.join(args.output_dir, f"{slug}_accrual_validation_{args.period_slug}.txt")

    # Build workpaper
    build_workpaper(
        matrix["vendors"], confirmed, completed_months,
        args.expense_acct_name, args.accrued_exp_acct_code,
        args.entity_id, args.period_end, args.period_slug, wp_path,
    )
    print(f"Workpaper: {wp_path}")

    # Build JE CSV
    total_debits, total_credits = build_je_csv(
        confirmed, args.expense_acct_code, args.expense_acct_name,
        args.accrued_exp_acct_code, args.period_slug, args.period_end, je_path,
    )
    print(f"JE CSV: {je_path}")
    print(f"Balance check: debits=${total_debits:,.2f} credits=${total_credits:,.2f} {'PASS' if abs(total_debits - total_credits) < 0.01 else 'FAIL'}")

    # Build validation
    build_validation(
        confirmed, excluded, completed_months, matrix["vendors"],
        total_debits, total_credits, args.expense_acct_name, args.period_slug, val_path,
    )
    print(f"Validation: {val_path}")


if __name__ == "__main__":
    main()
