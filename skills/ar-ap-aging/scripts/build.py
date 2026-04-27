"""
Build a lean AR or AP Aging workbook — 2 sheets, GL reconciliation, nothing else.

Usage: python build.py <config.json>

config.json:
{
    "mode": "AR",                # or "AP"
    "company_name": "Acme",
    "as_of_date": "2026-03-31",
    "gl_balance": 20290371.32,
    "output_path": "/abs/path/out.xlsx",
    "txn_tsv_path": "/abs/path/txns.tsv",
    "b2b_only": true             # AR only; ignored when mode == "AP"
}

TSV must include columns from Numeric's query_transaction_lines:
id, transaction_type, transaction_date, posting_date, normal_amount,
counterparty, organization_name, transaction_name

Uses `normal_amount` (oriented to the account's natural direction):
  AR — CustInvc positive; CustPymt / CustCred negative
  AP — VendBill positive; VendPymt / VendCred negative
"""
import csv
import json
import sys
from collections import defaultdict
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BUCKETS = ["Current", "1-30", "31-60", "61-90", "90+"]
EDGES = [30, 60, 90]

MODE_CONFIG = {
    "AR": {
        "trade_types": {"CustInvc", "CustPymt", "CustCred"},
        "open_type": "CustInvc",       # the item we age
        "cp_label": "Customer",
        "oldest_label": "Oldest Inv",
        "sheet_name": "Aging by Customer",
        "open_word": "invoices",
        "title_word": "AR Aging",
        "by_word": "by Customer",
    },
    "AP": {
        "trade_types": {"VendBill", "VendPymt", "VendCred"},
        "open_type": "VendBill",
        "cp_label": "Vendor",
        "oldest_label": "Oldest Bill",
        "sheet_name": "Aging by Vendor",
        "open_word": "bills",
        "title_word": "AP Aging",
        "by_word": "by Vendor",
    },
}


def parse_date(s):
    return date.fromisoformat(s[:10])


def bucket_idx(days):
    if days <= 0:
        return 0
    for i, edge in enumerate(EDGES):
        if days <= edge:
            return i + 1
    return len(EDGES) + 1


def load(tsv_path, as_of, trade_types, b2b_only):
    """Dedupe by id, filter to trade lines, drop post-as-of rows."""
    seen = set()
    trade = []
    with open(tsv_path) as f:
        for r in csv.DictReader(f, delimiter="\t"):
            if r["id"] in seen:
                continue
            seen.add(r["id"])
            try:
                if parse_date(r["posting_date"]) > as_of:
                    continue
                cp = (r["counterparty"] or "").strip()
                ttype = r["transaction_type"]
                if ttype not in trade_types:
                    continue  # Journal/FxReval/etc excluded
                if not cp:
                    continue  # empty counterparty = journal-ish, skip
                if b2b_only and cp.startswith("Cust:"):
                    continue  # AR DTC filter
                trade.append({
                    "cp": cp,
                    "type": ttype,
                    "txn_date": parse_date(r["transaction_date"]),
                    "amt": float(r["normal_amount"] or 0),
                    "entity": r.get("organization_name", ""),
                })
            except Exception:
                continue
    return trade


def fifo_age(trade, as_of, open_type):
    """Per-counterparty FIFO: sort open items, pool reductions, consume oldest first."""
    by_cp = defaultdict(list)
    for t in trade:
        by_cp[t["cp"]].append(t)

    result = {}
    for cp, items in by_cp.items():
        opens = sorted(
            [t for t in items if t["type"] == open_type and t["amt"] > 0],
            key=lambda x: x["txn_date"],
        )
        pool = abs(sum(t["amt"] for t in items if t["amt"] < 0))

        remaining = []
        for it in opens:
            if pool >= it["amt"]:
                pool -= it["amt"]
            elif pool > 0:
                remaining.append({"date": it["txn_date"], "open": it["amt"] - pool})
                pool = 0
            else:
                remaining.append({"date": it["txn_date"], "open": it["amt"]})

        aging = [0.0] * len(BUCKETS)
        oldest = None
        for it in remaining:
            days = (as_of - it["date"]).days
            aging[bucket_idx(days)] += it["open"]
            if oldest is None or it["date"] < oldest:
                oldest = it["date"]

        # Unapplied credits / overpayments → negative in Current
        if pool > 0:
            aging[0] -= pool

        total = sum(aging)
        if abs(total) > 0.01 or remaining:
            result[cp] = {
                "aging": aging,
                "total": total,
                "oldest": oldest.isoformat() if oldest else "",
                "entity": items[0]["entity"],
            }
    return result


# ---- Workbook ----

NAVY = PatternFill("solid", start_color="1F3864")
BLUE = PatternFill("solid", start_color="2F5597")
YELLOW = PatternFill("solid", start_color="FFF2CC")
GREY = PatternFill("solid", start_color="F2F2F2")
CCY = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
WHITE = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)


def build_summary(ws, cfg, result, mcfg):
    ws["A1"] = f"{cfg['company_name']} — {mcfg['title_word']} as of {cfg['as_of_date']}"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = NAVY
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 24

    ws["A2"] = f"Lookback: 12 months. Open {mcfg['open_word']} older than 12 months are not captured. Aged by transaction date."
    ws["A2"].font = Font(size=9, italic=True, color="595959")
    ws.merge_cells("A2:D2")

    totals = [0.0] * len(BUCKETS)
    for data in result.values():
        for i, v in enumerate(data["aging"]):
            totals[i] += v

    ws["A3"] = "Bucket"
    ws["B3"] = "Amount (USD)"
    ws["C3"] = "% of Observable"
    for col in "ABC":
        ws[f"{col}3"].font = WHITE
        ws[f"{col}3"].fill = BLUE
        ws[f"{col}3"].alignment = Alignment(horizontal="center")

    for i, b in enumerate(BUCKETS):
        r = 4 + i
        ws.cell(row=r, column=1, value=b)
        c = ws.cell(row=r, column=2, value=totals[i])
        c.number_format = CCY
        c = ws.cell(row=r, column=3, value=f"=B{r}/B{4+len(BUCKETS)}")
        c.number_format = "0.0%"

    tot_r = 4 + len(BUCKETS)
    ws.cell(row=tot_r, column=1, value="Total (observable)").font = BOLD
    c = ws.cell(row=tot_r, column=2, value=f"=SUM(B4:B{tot_r-1})")
    c.number_format = CCY
    c.font = BOLD
    for col in range(1, 4):
        ws.cell(row=tot_r, column=col).fill = GREY

    # GL reconciliation
    rec = tot_r + 2
    ws.cell(row=rec, column=1, value=f"GL trade {cfg['mode']} balance").font = BOLD
    c = ws.cell(row=rec, column=2, value=cfg["gl_balance"])
    c.number_format = CCY
    c.font = BOLD

    ws.cell(row=rec + 1, column=1, value="Less: observable aging")
    c = ws.cell(row=rec + 1, column=2, value=f"=B{tot_r}")
    c.number_format = CCY

    ws.cell(row=rec + 2, column=1, value="Unreconciled gap").font = BOLD
    c = ws.cell(row=rec + 2, column=2, value=f"=B{rec}-B{rec+1}")
    c.number_format = CCY
    c.font = BOLD
    ws.cell(row=rec + 2, column=1).fill = YELLOW
    ws.cell(row=rec + 2, column=2).fill = YELLOW

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16


def build_aging(ws, cfg, result, mcfg):
    ws["A1"] = f"{cfg['company_name']} — {mcfg['title_word']} {mcfg['by_word']} (FIFO)"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = NAVY
    ncols = 2 + len(BUCKETS) + 2
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.row_dimensions[1].height = 24

    hdr = [mcfg["cp_label"], "Entity"] + BUCKETS + ["Total", mcfg["oldest_label"]]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = WHITE
        c.fill = BLUE
        c.alignment = Alignment(horizontal="center")

    sorted_cp = sorted(result.items(), key=lambda x: -x[1]["total"])
    r = 4
    for cp, data in sorted_cp:
        ws.cell(row=r, column=1, value=cp)
        ws.cell(row=r, column=2, value=data["entity"])
        for i, v in enumerate(data["aging"]):
            c = ws.cell(row=r, column=3 + i, value=v)
            c.number_format = CCY
        first = get_column_letter(3)
        last = get_column_letter(2 + len(BUCKETS))
        c = ws.cell(row=r, column=3 + len(BUCKETS), value=f"=SUM({first}{r}:{last}{r})")
        c.number_format = CCY
        c.font = BOLD
        ws.cell(row=r, column=4 + len(BUCKETS), value=data["oldest"])
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = BOLD
    for i in range(len(BUCKETS) + 1):
        col = 3 + i
        L = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f"=SUM({L}4:{L}{r-1})")
        c.number_format = CCY
        c.font = BOLD
        ws.cell(row=r, column=col).fill = GREY
    ws.cell(row=r, column=1).fill = GREY
    ws.cell(row=r, column=2).fill = GREY

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 30
    for i in range(len(BUCKETS) + 1):
        ws.column_dimensions[get_column_letter(3 + i)].width = 14
    ws.column_dimensions[get_column_letter(ncols)].width = 14


def main():
    if len(sys.argv) != 2:
        print("Usage: build.py <config.json>")
        sys.exit(1)

    cfg = json.load(open(sys.argv[1]))
    mode = cfg.get("mode", "").upper()
    if mode not in MODE_CONFIG:
        print(f"Invalid mode: {mode!r}. Must be 'AR' or 'AP'.")
        sys.exit(1)
    cfg["mode"] = mode
    mcfg = MODE_CONFIG[mode]

    as_of = parse_date(cfg["as_of_date"])
    b2b_only = cfg.get("b2b_only", True) if mode == "AR" else False

    trade = load(cfg["txn_tsv_path"], as_of, mcfg["trade_types"], b2b_only)
    result = fifo_age(trade, as_of, mcfg["open_type"])
    print(f"Mode: {mode}  Trade lines: {len(trade)}  {mcfg['cp_label']}s with open balance: {len(result)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    build_summary(ws, cfg, result, mcfg)
    build_aging(wb.create_sheet(mcfg["sheet_name"]), cfg, result, mcfg)

    wb.save(cfg["output_path"])
    print(f"Saved: {cfg['output_path']}")


if __name__ == "__main__":
    main()
